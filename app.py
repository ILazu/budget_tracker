
import os
import io
from datetime import datetime
from typing import Tuple, List

import streamlit as st
import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.worksheet import Worksheet
import matplotlib.pyplot as plt
import qrcode

# -----------------------------
# CONFIG
# -----------------------------
APP_TITLE = "Desglose Económico Mensual"
FILE_NAME = "desglose_económico_esc_teo.xlsx"  # yearly file
SPANISH_MONTHS = [
    "", "Enero", "Febrero", "Marzo", "Abril", "Mayo", "Junio",
    "Julio", "Agosto", "Septiembre", "Octubre", "Noviembre", "Diciembre"
]


# -----------------------------
# UTILS (Excel I/O)
# -----------------------------
def month_sheet_name(year: int, month: int) -> str:
    return f"{SPANISH_MONTHS[month]} {year}"


def ensure_workbook(path: str):
    """Create workbook if not exists."""
    if not os.path.exists(path):
        wb = Workbook()
        # Create a minimal 'Inicio' sheet with instructions
        ws = wb.active
        ws.title = "Inicio"
        ws["A1"] = "Archivo creado por el Desglose Económico Mensual."
        ws["A2"] = "Las hojas se crearán automáticamente al ingresar datos por mes."
        wb.save(path)


def openpyxl_get_ws(wb, sheet_name: str) -> Worksheet:
    if sheet_name in wb.sheetnames:
        return wb[sheet_name]
    else:
        ws = wb.create_sheet(sheet_name)
        # Initialize headers for donations and expenses tables with a spacer row between them
        ws["A1"] = "DONACIONES"
        ws["A2"] = "Fecha"
        ws["B2"] = "Descripción"
        ws["C2"] = "Monto"
        # spacer row (row 3 and 4) to ensure clean separation
        ws["A3"] = None
        ws["A4"] = None
        # Expenses (Food & Snacks) table starting lower
        ws["A5"] = "GASTOS (Comida y Meriendas)"
        ws["A6"] = "Fecha"
        ws["B6"] = "Descripción"
        ws["C6"] = "Monto"
        return ws


def read_table(ws: Worksheet, start_row: int) -> pd.DataFrame:
    """Read a simple 3-column table (Fecha, Descripción, Monto) starting at start_row+1 (headers at start_row).
    Stops when it meets an empty row, a section title, or another header row to avoid bleeding into the next table.
    """
    rows = []
    row = start_row + 1
    # Known section titles and header row content
    SECTION_TITLES = {"DONACIONES", "GASTOS (Comida y Meriendas)"}
    HEADER_ROW = ("Fecha", "Descripción", "Monto")

    while True:
        c1 = ws.cell(row=row, column=1).value
        c2 = ws.cell(row=row, column=2).value
        c3 = ws.cell(row=row, column=3).value

        # Normalize to strings for comparisons without breaking numeric rows
        c1s = str(c1).strip() if c1 is not None else None
        c2s = str(c2).strip() if c2 is not None else None
        c3s = str(c3).strip() if c3 is not None else None

        # 1) stop if fully empty row
        if c1 is None and c2 is None and c3 is None:
            break
        # 2) stop if we hit the name of another section title in column A
        if c1s in SECTION_TITLES:
            break
        # 3) stop if we hit another header row (Fecha, Descripción, Monto)
        if (c1s, c2s, c3s) == HEADER_ROW:
            break

        rows.append([c1, c2, c3])
        row += 1
        # safety
        if row > 10000:
            break

    df = pd.DataFrame(rows, columns=["Fecha", "Descripción", "Monto"])
    # Normalize types
    if not df.empty:
        # Try to parse dates
        try:
            df["Fecha"] = pd.to_datetime(df["Fecha"]).dt.date
        except Exception:
            pass
        # Coerce amounts
        df["Monto"] = pd.to_numeric(df["Monto"], errors="coerce").fillna(0.0)
    return df


def append_row(ws: Worksheet, start_row: int, values: List):
    """Append a row to the first empty line after start_row+1 (headers at start_row)."""
    row = start_row + 1
    while True:
        c1 = ws.cell(row=row, column=1).value
        c2 = ws.cell(row=row, column=2).value
        c3 = ws.cell(row=row, column=3).value
        if c1 is None and c2 is None and c3 is None:
            break
        row += 1
        if row > 10000:
            raise RuntimeError("Tabla demasiado larga")
    for col, val in enumerate(values, start=1):
        ws.cell(row=row, column=col).value = val


def get_month_data(year: int, month: int) -> Tuple[pd.DataFrame, pd.DataFrame]:
    ensure_workbook(FILE_NAME)
    wb = load_workbook(FILE_NAME)
    ws = openpyxl_get_ws(wb, month_sheet_name(year, month))
    # Donations table starts at A2 (headers), i.e., start_row=2
    donations = read_table(ws, start_row=2)
    # Expenses table starts at A6 (headers), i.e., start_row=6
    expenses = read_table(ws, start_row=6)
    wb.save(FILE_NAME)
    return donations, expenses


def add_donation(year: int, month: int, date_str: str, desc: str, amount: float):
    ensure_workbook(FILE_NAME)
    wb = load_workbook(FILE_NAME)
    ws = openpyxl_get_ws(wb, month_sheet_name(year, month))
    append_row(ws, start_row=2, values=[date_str, desc, amount])
    wb.save(FILE_NAME)


def add_expense(year: int, month: int, date_str: str, desc: str, amount: float):
    ensure_workbook(FILE_NAME)
    wb = load_workbook(FILE_NAME)
    ws = openpyxl_get_ws(wb, month_sheet_name(year, month))
    append_row(ws, start_row=6, values=[date_str, desc, amount])
    wb.save(FILE_NAME)


def monthly_totals(don_df: pd.DataFrame, exp_df: pd.DataFrame) -> Tuple[float, float]:
    donations = float(don_df["Monto"].sum()) if not don_df.empty else 0.0
    expenses = float(exp_df["Monto"].sum()) if not exp_df.empty else 0.0
    return donations, expenses


def compute_previous_balance_for_month(year: int, month: int, initial_prev_jan: float) -> float:
    """
    Compute previous balance for given month by rolling from January to month-1:
    prev := initial_prev_jan
    for k in 1..month-1: prev = prev + donations_k - expenses_k
    """
    prev = float(initial_prev_jan)
    for m in range(1, month):
        don, exp = get_month_data(year, m)
        d, e = monthly_totals(don, exp)
        prev = prev + d - e
    return prev


# -----------------------------
# STREAMLIT UI
# -----------------------------
st.set_page_config(page_title=APP_TITLE, page_icon="💰", layout="wide")
st.title(APP_TITLE)

# Year & month selection
now = datetime.now()
col_y, col_m, col_mode = st.columns([1, 1, 1])
with col_y:
    year = st.number_input("Año", min_value=2000, max_value=2100, value=now.year, step=1)
with col_m:
    month = st.selectbox("Mes", list(range(1, 13)), index=now.month-1, format_func=lambda x: SPANISH_MONTHS[x])
with col_mode:
    viewer_mode = st.toggle("Modo solo lectura (para estudiantes)", value=True)

# Initial previous balance only used to start the year (January baseline)
with st.sidebar:
    st.header("Configuración")
    st.markdown("**Saldo inicial de enero** (solo si comienzas el año):")
    initial_prev_jan = st.number_input("Saldo previo de enero", min_value=0.0, step=50.0, value=0.0, format="%.2f")
    st.caption("Si ya registraste meses anteriores, este valor solo aplica si el archivo está vacío.")

    st.divider()
    st.subheader("Código de administrador (opcional)")
    admin_input = st.text_input("Introduce el código para habilitar edición", type="password")
    admin_secret = st.secrets.get("ADMIN_CODE", "")
    is_admin = bool(admin_secret) and admin_input == admin_secret
    if not admin_secret:
        st.info("Puedes establecer un **ADMIN_CODE** en *Secrets* al publicar en Streamlit Cloud para proteger la edición.")

    editing_enabled = (not viewer_mode) and (is_admin or not admin_secret)

# Data entry forms
st.subheader(f"Mes: {SPANISH_MONTHS[month]} {year}")
donations_df, expenses_df = get_month_data(year, month)

col1, col2 = st.columns(2)
with col1:
    st.markdown("### Registrar Donación")
    with st.form("donation_form", clear_on_submit=True):
        d_date = st.date_input("Fecha", value=datetime.now().date())
        d_desc = st.text_input("Descripción (opcional)", placeholder="Donativo anónimo / actividad X")
        d_amount = st.number_input("Monto", min_value=0.0, step=1.0, format="%.2f")
        submitted_d = st.form_submit_button("Agregar donación", disabled=not editing_enabled)
        if submitted_d:
            add_donation(year, month, d_date.isoformat(), d_desc, float(d_amount))
            st.success("Donación registrada.")
            st.rerun()

with col2:
    st.markdown("### Registrar Gasto (Comida y Meriendas)")
    with st.form("expense_form", clear_on_submit=True):
        e_date = st.date_input("Fecha ", value=datetime.now().date(), key="ex_date")
        e_desc = st.text_input("Descripción (opcional)", placeholder="Pizza reunión / meriendas asamblea", key="ex_desc")
        e_amount = st.number_input("Monto ", min_value=0.0, step=1.0, format="%.2f", key="ex_amt")
        submitted_e = st.form_submit_button("Agregar gasto", disabled=not editing_enabled)
        if submitted_e:
            add_expense(year, month, e_date.isoformat(), e_desc, float(e_amount))
            st.success("Gasto registrado.")
            st.rerun()

st.divider()

# Compute rolling previous balance
prev_balance = compute_previous_balance_for_month(year, month, initial_prev_jan)
don_total, exp_total = monthly_totals(donations_df, expenses_df)
total_budget = prev_balance + don_total
remaining = total_budget - exp_total

left, right = st.columns(2)
with left:
    st.markdown("### Resumen del mes")
    st.metric("Saldo previo", f"${prev_balance:,.2f}")
    st.metric("Donaciones", f"${don_total:,.2f}")
    st.metric("Presupuesto total (Previo + Donaciones)", f"${total_budget:,.2f}")
with right:
    st.markdown("### Estado de gastos")
    st.metric("Gastos (Comida y Meriendas)", f"${exp_total:,.2f}")
    st.metric("Saldo restante", f"${remaining:,.2f}")

# Tables
st.markdown("#### Donaciones")
st.dataframe(donations_df if not donations_df.empty else pd.DataFrame(columns=["Fecha", "Descripción", "Monto"]), use_container_width=True)
st.markdown("#### Gastos (Comida y Meriendas)")
st.dataframe(expenses_df if not expenses_df.empty else pd.DataFrame(columns=["Fecha", "Descripción", "Monto"]), use_container_width=True)

# Charts
st.divider()
st.subheader("Gráficas")

# Pie 1: Previous balance vs donations
fig1 = plt.figure(figsize=(4, 4))
plt.pie([prev_balance, don_total], labels=["Saldo Previo", "Donaciones"], autopct='%1.1f%%', startangle=90)
plt.title("Origen de fondos")
st.pyplot(fig1)

# Pie 2: Total budget vs amount spent (and remaining)
fig2 = plt.figure(figsize=(4, 4))
plt.pie([exp_total, max(remaining, 0.0)], labels=["Gastado", "Restante"], autopct='%1.1f%%', startangle=90)
plt.title("Uso del presupuesto")
st.pyplot(fig2)

# Download Excel
if os.path.exists(FILE_NAME):
    with open(FILE_NAME, "rb") as f:
        st.download_button("Descargar Excel del año", data=f, file_name=FILE_NAME, mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

# QR generator for public viewing link
st.divider()
st.subheader("Código QR para compartir (solo visualización)")
st.caption("Introduce la **URL pública** de este dashboard (p. ej., la URL de Streamlit Cloud) para generar un QR que tus compañeros puedan escanear.")
public_url = st.text_input("URL pública del dashboard", placeholder="https://tu-app-streamlit.streamlit.app/")

col_qr1, col_qr2 = st.columns([1, 2])
with col_qr1:
    if st.button("Generar QR") and public_url.strip():
        qr_img = qrcode.make(public_url.strip())
        buf = io.BytesIO()
        qr_img.save(buf, format="PNG")
        st.session_state["qr_bytes"] = buf.getvalue()

    if "qr_bytes" in st.session_state:
        st.image(st.session_state["qr_bytes"], caption="Escanea para abrir el dashboard")
        st.download_button("Descargar QR", data=st.session_state["qr_bytes"], file_name="qr_dashboard.png", mime="image/png")

with col_qr2:
    st.info("Publica esta app en **Streamlit Community Cloud**, configura `ADMIN_CODE` en *Secrets* para proteger la edición, y comparte este QR con los estudiantes para que vean tablas y gráficas en modo lectura.")

st.caption("© Student Council Budget • Construido con Streamlit, OpenPyXL, Matplotlib y QRCode")
